import importlib
import subprocess
import sys

# Liste des bibliothèques nécessaires
required_libs = [
    "pandas",
    "openpyxl",
    "os",
    "calendar",
    "urllib.parse",
    "requests",
    "urllib3",
    "datetime",
    "hashlib",
    "time",
    "matplotlib",
    "math",
    "pvlib"
]

# Fonction pour vérifier et installer les bibliothèques manquantes
def check_and_install(package_name):
    try:
        importlib.import_module(package_name.split('.')[0])
    except ImportError:
        print(f"Installation de {package_name}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])

# Vérifie toutes les dépendances
for lib in required_libs:
    check_and_install(lib)


import pandas as pd
import os

import calendar
import urllib.parse
import requests
import urllib3
from datetime import datetime
import hashlib
import time
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

from matplotlib.colors import LinearSegmentedColormap  # Ajout de l'importation manquante

from pvlib import solarposition


import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.collections import LineCollection
from datetime import  timedelta
import math

from openpyxl.drawing.image import Image





urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

############################################################################################"
#USER DATA

doaall=0;   # mettre a un pour ignorer les fichiers existants et tout refaire

# Variables globales
resultats = []
lat, lon = 45.762, 4.698
startyear, endyear = 2019, 2023

prix_kw = 0.2016  # €/kWh
date_departcumul = pd.Timestamp("2019-01-01")
conso_chauffeau_journalière = 2.2  # kWh

conso_maison_W= 100
# Définition des masks d'horizon

#attention si les data de l'horizon change il faut efface les fichiers intermediare ( faire doaall=1)  pour refaire la demande a pvgis
## definir un profile par lieu de panneau ( hauteur des masques de 0 (nord) a 360 degree )



datahorizonType=[
    "" ,
    #1:balcon
     "4.7266,4.7253,4.7241,4.7228,4.7215,4.7203,4.719,4.7178,4.7165,4.7153,4.714,4.7127,4.7115,4.7102,4.709,4.7077,4.7065,4.7052,\
4.704,4.7027,4.7014,4.7002,4.6989,4.6977,4.6964,4.6952,4.6939,4.6926,4.6914,4.6901,4.6889,4.6876,4.6864,4.6851,4.6839,\
4.6826,4.6813,4.6801,4.6788,4.6776,4.6763,4.6751,4.6738,4.6725,4.6713,4.67,4.6688,4.6675,4.6663,4.665,4.6638,4.6625,\
4.6612,4.66,4.6587,4.6575,4.6562,4.655,4.6537,4.6524,4.6512,4.6499,4.6487,4.6474,4.6462,4.6449,4.6437,4.6424,4.6411,\
4.6399,4.6386,4.6374,4.6364,4.6364,4.596,4.4343,3.9899,5.7273,5.9697,7.7072,10.9577,16.7717,18.4817,21.5892,24.2871,\
25.4848,25.5657,25.3232,24.8788,24.5152,24.0707,23.6667,22.9838,22.4949,22.1981,21.9871,21.202,20.555,21.3924,22.7156,\
25.8357,25.8523,25.2424,24.9454,25.2484,26.6162,27.7887,28.5556,27.9711,21.0061,15.8723,14.6566,15.4791,16.3365,17.5901,\
19.5455,19.9495,20.413,21.202,21.5623,22.1369,21.3086,20.6768,20.6286,21.2828,20.6768,19.0625,17.5778,17,16.3535,15.5306,\
14.956,13.8964,10.9726,9.404,8.2144,6.6294,3.6726,3.1414,4.3282,7.8196,9.382,10.6582,11.1521,11.4242,10.8299,10.4114,\
12.506,14.7329,16.095,20.1776,22.214,23.5455,24.2522,25.0289,25.6061,26.2994,27.4148,28.3131,29.2355,29.2828,29.2828,\
29.5253,29.8081,29.7273,29.6869,29.6465,29.6307,30.6903,32.1538,33.8016,35.303,35.6152,35.907,35.9091,35.8528,34.47,\
32.21,28.4268,25.837,24.3125,22.4064,20.7999,18.6756,11.3115,4.2378,1.6402,1.6869,1.7677,1.9697,1.7677,1.7677,1.8154,\
1.8485,1.9697,2.4485,2.4112,2.2525,2.4351,2.7374,2.7374,2.6162,2.899,3.0202,3.6263,3.8251,3.3434,3.2222,3.2524,3.3838,\
3.1414,3.1414,3.652,3.5051,3.4242,3.3838,3.4646,4.4243,5.9697,4.4747,4.2154,3.9899,3.8687,4.1515,4.1515,3.9495,3.8283,\
3.8687,4.0303,4.1111,4.1111,4.3939,4.4747,4.5152,4.5152,4.5132,4.4747,4.7172,4.8384,5.202,5.202,5.206,5.4444,5.404,5.404,\
5.6461,5.7677,5.7677,5.6061,5.4848,5.5657,5.7677,5.4848,5.5253,5.8081,6.0593,7.4242,7.101,7.5721,7.267,8.2995,6.899,6.4141,\
6.4949,8.6454,8.4567,6.7778,7.7475,7.7521,8.5288,9.2828,8.7664,8.1917,7.7787,6.5758,9.0272,9.5253,8.9596,8.1653,8.7804,\
7.9177,9.1212,9.6653,9.2424,8.4069,7.6706,7.5051,7.5051,8.1919,8.2323,8.1919,8.3939,7.7475,7.8559,8.1919,8.1164,8.5746,\
6.899,5.9537,5.2012,5.1616,5.1616,4.8384,4.7969,4.7957,4.7944,4.7932,4.7919,4.7906,4.7894,4.7881,4.7869,4.7856,4.7844,\
4.7831,4.7818,4.7806,4.7793,4.7781,4.7768,4.7756,4.7743,4.7731,4.7718,4.7705,4.7693,4.768,4.7668,4.7655,4.7643,4.763,\
4.7617,4.7605,4.7592,4.758,4.7567,4.7555,4.7542,4.753,4.7517,4.7504,4.7492,4.7479,4.7467,4.7454,4.7442,4.7429,4.7416,\
4.7404,4.7391,4.7379,4.7366,4.7354,4.7341,4.7328,4.7316,4.7303,4.7291,4.7278,4.7266",

    #2:portail


"34.6957,34.6572,34.6186,34.5801,34.5416,34.5030,34.4645,34.4260,34.3874,34.3489,34.3103,34.2718,34.2333,34.1947,34.1562,\
34.1177,34.0791,34.0406,34.0021,33.9635,33.9250,33.8864,33.8479,34.0636,34.2807,34.6912,35.0927,35.4035,35.7203,36.2456,\
36.3128,36.5965,36.9053,37.1314,37.3684,37.3684,37.4386,37.5789,37.5088,37.4737,37.3684,37.3407,37.1228,36.9474,36.9123,\
36.8772,36.8773,36.9474,36.9825,37.1228,37.2202,37.4737,37.3803,37.4774,37.6684,38.1050,38.4561,38.7018,38.9474,39.1579,\
39.4035,39.6719,39.8947,40.0702,40.0351,40.0000,40.0000,39.9298,39.6578,39.3684,39.3684,39.1228,38.7185,38.3860,38.0909,\
37.5789,36.8668,36.0702,35.4322,34.7719,33.8947,33.0698,32.2121,27.6842,23.7895,21.5339,20.2201,19.2221,18.3294,17.7193,\
17.0526,16.5985,15.9028,14.6316,11.9554,8.8872,7.9298,7.9649,8.0702,8.0351,7.9649,8.0262,8.2456,7.7544,7.5685,8.4561,8.3158,\
5.9671,7.8947,7.5789,5.5088,7.0526,6.8070,6.0000,12.3509,12.9123,13.6849,14.2618,14.6667,13.9054,13.7544,13.8865,13.2642,\
13.5396,12.5263,13.0877,12.0000,12.2143,10.1286,10.3080,10.3158,10.6316,10.8421,11.6756,14.7719,14.7018,14.2190,12.0000,\
11.6222,11.0102,10.5034,10.0351,9.8947,9.4737,8.8772,8.3509,10.8506,12.6229,15.2419,16.5390,15.6111,13.6491,10.9210,5.3374,\
1.5088,1.7753,3.0683,6.6030,7.3203,7.8972,8.3860,8.7719,9.1719,9.5734,9.9649,10.0351,10.0000,10.0000,9.7895,9.4210,9.0526,\
8.7018,8.1754,7.6842,7.2029,7.1172,6.9825,7.1595,8.5778,9.8270,11.1228,11.7895,12.2548,13.0422,14.0702,15.6795,15.9333,\
17.5088,17.5789,17.4386,16.5263,16.9386,17.1471,17.3333,17.5992,18.0000,18.2456,18.4243,17.4386,16.8421,19.9649,19.9411,\
19.4035,16.6468,15.7532,14.5614,14.7368,14.9123,14.9474,14.9474,15.0526,15.1617,15.3333,15.3684,15.3684,15.3684,15.4737,\
15.5789,15.6140,15.6842,15.5789,10.1499,4.1058,4.1754,4.2853,5.1689,5.9914,6.7438,6.6091,5.8918,3.9795,2.2456,2.5263,2.3158,\
2.1860,2.9474,2.8266,2.6446,2.6692,3.7895,3.8582,3.2071,2.7018,3.3434,3.7287,3.4447,3.0661,2.8421,2.9123,3.2964,3.7680,\
4.1404,4.7018,6.2313,10.2049,9.0804,7.2598,5.4769,3.3333,3.1930,3.0877,2.7996,3.3932,4.1404,5.2345,6.1767,7.1228,8.1699,\
9.2058,10.1404,11.1667,12.2700,13.3429,14.5467,15.4745,16.4725,17.2600,18.0825,19.0526,20.0518,21.0303,22.2543,23.2831,\
24.2456,24.7878,25.8488,27.5326,29.1503,30.6044,31.6726,32.2807,32.6316,32.8070,33.0330,33.1889,33.2982,33.5168,33.6842,\
33.7193,33.9298,34.1945,34.2456,34.4862,34.6667,34.9474,34.9474,34.9474,35.0877,35.4386,35.4386,35.5439,35.5585,35.6140,\
35.7544,36.1314,36.3913,36.3528,36.3142,36.2757,36.2372,36.1986,36.1601,36.1215,36.0830,36.0445,36.0059,35.9674,35.9289,\
35.8903,35.8518,35.8133,35.7747,35.7362,35.6976,35.6591,35.6206,35.5820,35.5435,35.5050,35.4664,35.4279,35.3894,35.3508,\
35.3123,35.2738,35.2352,35.1967,35.1581,35.1196,35.0811,35.0425,35.0040,34.9655,34.9269,34.8884,34.8499,34.8113,34.7728,\
34.7342,34.6957",
    
  #3 bas terrain  
"9.9652,10.1404,10.3156,10.4907,10.6659,10.8411,11.0162,11.1914,11.3666,11.5417,11.7169,11.8921,12.0672,12.2424,12.4176,\
12.5927,12.7679,12.9430,13.1182,13.2934,13.4685,13.6437,13.8189,13.9940,14.1692,14.3444,14.5195,14.6947,14.8699,15.0450,\
15.2202,15.3954,15.5705,15.7457,15.9209,16.0960,16.2712,16.4464,16.6215,16.7967,16.9719,17.1470,17.3222,17.4974,17.6725,\
17.8477,18.0229,18.1980,18.3732,18.5483,18.7235,18.8987,19.0738,19.2490,19.4242,19.5993,19.7745,19.9497,20.1248,20.3000,\
20.4752,20.6503,20.8255,21.0007,21.1758,21.3510,21.5262,21.6788,21.7883,21.8978,22.0073,22.1533,22.2628,22.3723,22.4818,\
22.5912,22.7372,25.1643,26.1314,23.2117,23.3212,23.4307,22.5912,22.7372,22.8832,23.0292,23.1752,23.3212,23.4672,23.6131,\
22.5299,19.4293,16.3287,8.1387,8.1022,9.5247,10.7837,10.8029,9.3836,8.0292,8.4672,8.6278,8.9192,17.5912,17.8832,18.6496,\
19.2336,19.1606,18.4562,16.7779,15.9025,15.4745,15.4745,14.8457,12.2263,19.4452,22.1736,24.1731,25.0000,24.3327,21.4094,\
16.0336,13.3140,11.3838,11.1679,11.3524,12.3023,14.5630,16.0223,20.8006,26.6752,30.7653,33.9764,34.7788,35.0703,35.6934,\
36.2006,36.7883,37.2263,37.8102,38.2423,38.4973,38.6131,38.6131,38.6131,38.5766,38.3757,37.9562,37.6277,37.0073,36.6900,\
36.2044,35.3041,34.8540,33.8453,32.6780,31.6201,30.5987,29.7963,28.5925,27.0967,25.6373,24.4700,22.6086,19.3980,12.7927,\
11.4064,10.6569,11.0949,11.9862,12.7521,13.9559,14.7583,15.2920,15.4507,15.5109,15.5109,15.5109,15.3650,15.2190,14.6715,\
13.8960,13.1301,12.2547,11.0874,10.3650,10.3285,10.5839,10.9124,10.9124,10.7664,10.6204,10.6569,11.3504,12.0095,12.8102,\
14.9671,16.4970,17.9198,18.9047,19.7436,20.5095,20.9124,21.0219,21.1314,21.0219,21.0219,20.9124,20.7664,20.6204,20.5839,\
20.2920,19.7810,19.1683,18.6131,18.1110,17.3544,16.3967,15.0937,13.5255,11.5922,8.9051,8.2482,7.8832,7.7372,7.2768,7.0803,\
6.8248,7.3254,10.4786,11.2409,10.5647,8.4661,7.2993,7.4265,8.3019,9.1407,10.6366,11.4755,11.8764,12.2628,12.3947,12.2993,\
12.6277,12.6277,12.6277,12.6277,12.6642,12.7372,12.7372,12.5182,11.9620,11.4151,10.9412,10.3650,9.8175,9.5985,9.4161,9.0146,\
8.6451,8.2441,7.7007,7.4423,6.5328,6.0584,5.8015,5.5109,5.5109,5.6204,6.0212,2.5474,6.0219,5.9854,5.6569,6.6755,8.2577,9.5942,\
11.1314,10.5839,10.1825,12.9497,16.4163,15.7010,16.9182,18.6775,18.8321,18.7956,18.5910,19.4267,17.6390,16.7271,14.7203,13.6861,\
19.0146,17.8241,17.0073,17.7372,17.3723,17.3358,16.8613,17.0438,16.8613,16.8978,17.1354,16.3330,16.2409,16.1314,16.1314,15.5169,\
14.3796,13.6131,13.7591,13.7591,13.7591,12.6156,13.2598,13.0657,13.1022,11.8412,10.8664,10.3560,9.8905,9.4383,7.2554,7.7668,7.6579,\
6.3080,5.4015,5.1412,5.1460,4.8905,5.0606,5.2358,5.4109,5.5861,5.7613,5.9364,6.1116,6.2868,6.4619,6.6371,6.8123,6.9874,7.1626,\
7.3378,7.5129,7.6881,7.8632,8.0384,8.2136,8.3887,8.5639,8.7391,8.9142,9.0894,9.2646,9.4397,9.6149,9.7901,9.9652"


]





    
# Définition des scénarios de pose des panneaux

#note:
 # pas d'espace dans les libellé
 # tracking    0= fixe   2 2 axes    3 axe vertical qui tourne      5 axe incliné qui tourne
 # horizonType 0 = use default location   autre indice dans la table des profils d'horizon


installation = 120     # cable ,tableau elec , routeur solaire a triac bricolé
fixationPanneau=20
PV=80
panneau = PV + fixationPanneau
microonduleur1voie=126
microonduleur2voie=190
microonduleur4voie=250
tracker = 100


pourunpanneau=installation+panneau+microonduleur1voie
pourdeuxpanneaux=installation+panneau+panneau+microonduleur2voie

pourdeuxpanneauxsepare =installation+ panneau+ microonduleur1voie +panneau+microonduleur1voie

scenarios = [

            
    [         
        {"libelleScenario": "1xPortail", "investissement": pourunpanneau ,"conso_maison_W":conso_maison_W},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2 , "libelle": "optimal"},
    ],
    
            [     
        {"libelleScenario": "x1Tracker2axes_balcon", "investissement": pourunpanneau + tracker ,"conso_maison_W":conso_maison_W },
        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 2 , "horizonType": 1 , "libelle": "trackerbalcon"},
    ],

            [     
        {"libelleScenario": "x1Tracker2axes_basterrain", "investissement": pourunpanneau + tracker ,"conso_maison_W":conso_maison_W },
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 2 , "horizonType": 3 , "libelle": "trackerbasterrain"},
    ],

     [  
        {"libelleScenario": "1xBasterrain", "investissement": pourunpanneau ,"conso_maison_W":conso_maison_W},
         {"angle": 37, "aspect": 0 , "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 3, "libelle": "basterrainsud"},
     ],

     [  
        {"libelleScenario": "2xBasterrain", "investissement": pourdeuxpanneaux,"conso_maison_W":conso_maison_W},
         {"angle": 37, "aspect": 6 , "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 3, "libelle": "basterrainsud"},
         {"angle": 37, "aspect": 6 , "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 3, "libelle": "basterrainsud"},
     ],
            
     [  
        {"libelleScenario": "1xToitureEst_1xBalcon", "investissement": pourdeuxpanneaux,"conso_maison_W":conso_maison_W},
         {"angle": 23, "aspect": -(180-72) , "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "est"},
         {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1 , "libelle": "balcon_72"},
 
    ],
     
    [  
        {"libelleScenario": "2xToitureOuest_sansmask", "investissement": pourdeuxpanneaux,"conso_maison_W":conso_maison_W},
        {"angle": 23, "aspect": 72, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "ouestSM"},
        {"angle": 23, "aspect": 72, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "ouestSM"},
    ],

    [  
        {"libelleScenario": "1xToitureEst_1xToitureouest_sansmask", "investissement": pourdeuxpanneaux,"conso_maison_W":conso_maison_W},
        {"angle": 23, "aspect": 72, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "ouest"},
        {"angle": 23, "aspect": -(180-72) , "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "est"},
    ],



     [  
        {"libelleScenario": "1xBalcon_1xBalconOuest", "investissement": pourdeuxpanneaux,"conso_maison_W":conso_maison_W},
         {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1 , "libelle": "balcon_72"},
         {"angle": 72, "aspect": 72, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "ouest72"},
 
    ],
 

                        [     
        {"libelleScenario": "x1Tracker2axes_portail", "investissement": pourunpanneau + tracker ,"conso_maison_W":conso_maison_W },
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 2 , "horizonType": 2 , "libelle": "tracker"},
    ],

                        [     
        {"libelleScenario": "x2Tracker_2axes_Portail", "investissement": installation+panneau+panneau+tracker+tracker+microonduleur2voie  ,"conso_maison_W":conso_maison_W },
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 2 , "horizonType": 2 , "libelle": "tracker"},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 2 , "horizonType": 2 , "libelle": "tracker"},
    ],
            


    [  
        {"libelleScenario": "1xBalcon", "investissement": pourunpanneau ,"conso_maison_W":conso_maison_W },
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1 , "libelle": "balcon_72"},
    ],




    [  
        {"libelleScenario": "2xBalcon", "investissement": pourdeuxpanneaux,"conso_maison_W":conso_maison_W},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
    ],
    [  
        {"libelleScenario": "2xGarageOuest_sansmask", "investissement": pourdeuxpanneaux,"conso_maison_W":conso_maison_W},
        {"angle": 23, "aspect": 47, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "garage_ouestsm"},
        {"angle": 23, "aspect": 47, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "garage_ouestsm"},
    ],
    [  # mix optimal + balcon 
        {"libelleScenario": "1xPortail_1xBalcon", "investissement": pourdeuxpanneauxsepare,"conso_maison_W":conso_maison_W},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimalportail"},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
    ],
    [  # optimal x2  
        {"libelleScenario": "2xPortail", "investissement": pourdeuxpanneaux,"conso_maison_W":conso_maison_W},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimalportail"},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimalportail"},
    ],


        [  #  mix 
        {"libelleScenario": "1xPortail_2xBalcon", "investissement": installation + 3*panneau+microonduleur2voie + microonduleur1voie ,"conso_maison_W":conso_maison_W},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimalportail"},
 
    ],

        [  #  mix 
        {"libelleScenario": "2xPortail_1xBalcon", "investissement": installation + 3*panneau+ microonduleur2voie + microonduleur1voie,"conso_maison_W":conso_maison_W},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 3, "libelle": "optimalbasterrainsud"},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 3, "libelle": "optimalbasterrainsud"},
 
    ],

            
        [  #  mix 
        {"libelleScenario": "2xPortail_2xBalcon", "investissement": installation + 4*panneau+ 2* microonduleur2voie  ,"conso_maison_W":conso_maison_W},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimalportail"},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimalportail"},
  
    ],


        [  #  mix rigolo
        {"libelleScenario": "3xPortail_look", "investissement": installation + 3*panneau+microonduleur4voie  ,"conso_maison_W":conso_maison_W},
        {"angle": 89, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "portailvertical"},
        {"angle": 89, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "portailvertical"},
        {"angle": 37, "aspect": 0, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimalportail"},
 
    ],
    
  
]

#END USER DATA
############################################################################################"





def build_pvgis_url(i,lat, lon, angle, aspect,libelle, startyear, endyear, TrackerType, horizonType ,  MaxPower=1.00, tech='crystSi', db='PVGIS-SARAH3'):
    base_url = "https://re.jrc.ec.europa.eu/api/v5_3/seriescalc"
    params = {
        "lat": lat,
        "lon": lon,
        "raddatabase": db,
        "outputformat": "csv",
        "usehorizon": 1,
        "angle": angle,
        "aspect": aspect,
        "startyear": startyear,
        "endyear": endyear,
        "mountingplace": "free",
        "trackingtype": TrackerType,              #0= fixe   2 2 axes    3 axe vertical qui tourne      5 axe incliné qui tourne 
        "pvcalculation": 1,
        "pvtechchoice": tech,
        "peakpower": MaxPower,
        "loss": 14,
        "select_database_hourly":"PVGIS-SARAH3",

        
    }
    if horizonType != 0 :
        params["userhorizon"] = datahorizonType[horizonType]

        
    url = base_url + "?" + urllib.parse.urlencode(params)


    filename = f"_{libelle}_From{startyear}_To{endyear}_AN{params['angle']}_AZ{params['aspect']}_PW{MaxPower}_TR{TrackerType}_HO{horizonType}"



 #   filename = f"idata{i}_{angle}deg_{aspect}deg_{startyear}_{endyear}_PMax{MaxPower:.2f}_tracking{TrackerType}_hor{horizonType}.csv"
    return url, filename


def telecharger_csv(url, filename):
    if not os.path.exists(filename)  or doaall==1:
        print(f"Téléchargement depuis : {url}")
        response = requests.get(url, verify=False)
        if response.ok:
            with open(filename, 'wb') as f:
                f.write(response.content)
            print(f"✅ Fichier enregistré sous : {filename}")
            return 1
        else:
            print(f"❌ Erreur {response.status_code}")
            return 1
    else:
        print(f"✔️ Fichier déjà existant : {filename}")
        return 0


def traitement_csv(filename, maxpower):
    with open(filename, 'r', encoding='utf-8') as f:
        lignes = f.readlines()
    for i, ligne in enumerate(lignes):
        if ligne.strip().startswith("time,P"):
            header_idx = i
            break

    df = pd.read_csv(filename, skiprows=header_idx)
    df.columns = df.columns.str.strip().str.lower()
    df[['date', 'time']] = df['time'].str.split(":", expand=True)
    df = df[df['date'].astype(str).str.match(r'^\d{8}$')]
    df['date'] = pd.to_datetime(df['date'], format='%Y%m%d')
    df['time'] = df['time'].astype(str).str.zfill(4)
    df['hour'] = df['time'].str[:-2].astype(int)
    df['minute'] = df['time'].str[-2:].astype(int)
    df['datetime'] = df['date'] + pd.to_timedelta(df['hour'], unit='h') + pd.to_timedelta(df['minute'], unit='m')
    df['p'] = pd.to_numeric(df['p'], errors='coerce')

    return df






def traitement_cumul(filename,conso_maison_W):

    df = pd.read_excel(filename)
    conso_maison_kW = conso_maison_W / 1000
   
       
    df['date'] = pd.to_datetime(df['datetime']).dt.date
   
    df['p_dispo'] = 0.0
    df['p_maison'] = 0.0
    df['p_chauffeau'] = 0.0
    df['p_perdue'] = 0.0

    df['p_besoin'] = 0.0
    usableMaison=12  #arbitraire 12 heures de conso maison/j, div par 2 pour tenir compte nuit ou on ne pourra jamais recuperer
    for date_jour, df_jour in df.groupby(df['datetime'].dt.date):
        energie_chauffeau_cumul = 0.0
        for i, row in df_jour.iterrows():
            p_dispo = row['p']
            df.at[i, 'p_dispo']=p_dispo
            if p_dispo >= conso_maison_W:
                df.at[i, 'p_maison'] = conso_maison_W
                p_rest = p_dispo - conso_maison_W
            else:
                df.at[i, 'p_maison'] = p_dispo
                p_rest = 0

            if energie_chauffeau_cumul < conso_chauffeau_journalière:
                reste = (conso_chauffeau_journalière - energie_chauffeau_cumul) * 1000
                chauffe_eau = min(p_rest, reste)
                df.at[i, 'p_chauffeau'] = chauffe_eau
                energie_chauffeau_cumul += chauffe_eau / 1000
                p_rest -= chauffe_eau

            df.at[i, 'p_perdue'] = p_rest

            df.at[i,  'p_besoin'] =    conso_chauffeau_journalière*1000 /24  + conso_maison_W / (24/usableMaison)  


    df['date'] = pd.to_datetime(df['datetime'].dt.date)

    # Étape 1 : Colonnes d'agrégation
    scenario_cols = [col for col in df.columns if col.startswith('scenario_')]
    other_cols = ['p_besoin', 'p_dispo', 'p_maison', 'p_chauffeau', 'p_perdue']

    agg_dict = {col: 'sum' for col in scenario_cols + other_cols}

    # Étape 2 : Groupby somme
    daily = df.groupby('date').agg(agg_dict) / 1000

    # Étape 3 : Ajout du max de p_chauffeau (après groupby)
    daily['p_chauffeau_max'] = df.groupby('date')['p_chauffeau'].max() / 1000

    # Étape 4 : Calcul du manque
    daily['manque'] = (
        (conso_chauffeau_journalière + conso_maison_kW * usableMaison) -
        (daily['p_chauffeau'] + daily['p_maison'])
    ).clip(lower=0)


    # Ajustement: si 'energie_perdue' > 0, alors 'manque' = 0
  #  daily.loc[daily['p_perdue'] > 0, 'manque'] = 0

    daily = daily.rename(columns={
      
        'p_besoin': 'besoinKwh',
        'p_maison': 'energie_maison',
        'p_chauffeau': 'energie_chauffeau',
        'p_perdue': 'energie_perdue'
    })
    daily['energie_recuperee'] = daily['energie_maison'] + daily['energie_chauffeau']



    daily['gain (€)'] = daily['energie_recuperee'] * prix_kw
    daily['perte (€)'] = daily['energie_perdue'] * prix_kw
    daily['manque (€)'] = daily['manque'] * prix_kw

    daily['cumul_gain (€)'] = daily[daily.index >= date_departcumul]['gain (€)'].cumsum()
    daily['cumul_perte (€)'] = daily[daily.index >= date_departcumul]['perte (€)'].cumsum()

    daily['cumul_manque (€)'] = daily[daily.index >= date_departcumul]['manque (€)'].cumsum()

    daily['cumul_besoin (Kwh)'] = daily[daily.index >= date_departcumul]['besoinKwh'].cumsum()


    daily.loc[daily.index < date_departcumul, ['cumul_gain (€)', 'cumul_perte (€)']] = float('nan')

    
    return daily




def ajouter_graphique_excel(fichier_excel, df_resultat):
    wb = load_workbook(fichier_excel)
    ws = wb.active

    # Recherche de l’index (colonne) de 'energie_maison'
    header = [cell.value for cell in ws[1]]
    try:
        col_energie_maison = header.index("energie_maison") + 1  # Excel est 1-based
    except ValueError:
        print("⚠️ Colonne 'energie_maison' non trouvée dans le fichier Excel.")
        return

    # Conversion des dates pour grouper par mois
    df_resultat["date"] = pd.to_datetime(df_resultat.index)
    df_resultat["mois"] = df_resultat["date"].dt.to_period("M")

    start_row = 2
    row_offset = 10
    chart_height = 5
    chart_width = 20

    for i, mois in enumerate(df_resultat["mois"].unique()):
        df_mois = df_resultat[df_resultat["mois"] == mois]
        excel_start = df_resultat.index.get_loc(df_mois.index[0]) + 2
        excel_end = excel_start + len(df_mois) - 1

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = f"{calendar.month_name[mois.month]} {mois.year}"
        chart.y_axis.title = "Énergie (kWh)"
        chart.x_axis.title = "Jour"

        # Utilise la colonne 'energie_maison' et celle juste après (si tu veux plusieurs)
        data = Reference(ws,
                         min_col=col_energie_maison,
                         max_col=col_energie_maison + 3,  # Ajuste selon besoin
                         min_row=excel_start - 1,
                         max_row=excel_end)

        # Catégories (date/jour)
        cats = Reference(ws, min_col=1, max_col=1, min_row=excel_start, max_row=excel_end)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = chart_height
        chart.width = chart_width

        ws.add_chart(chart, f"Q{start_row + i * row_offset}")

    wb.save(fichier_excel)




def format_excel(excelfilename):
       
    # Ouvre le fichier pour modifications avec openpyxl
    wb = load_workbook(excelfilename)
    ws = wb.active

    # 1. Ajoute une ligne de filtrage (auto-filter)
    ws.auto_filter.ref = ws.dimensions  # Applique le filtre sur toute la plage de données

    # 2. Ajuste la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Numéro de colonne
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2  # +2 pour un petit confort visuel

    # 3. Souligne en bleu la dernière colonne si c'est un lien
    last_col_idx = ws.max_column
    for row in range(2, ws.max_row + 1):  # Ignore l'entête
        cell = ws.cell(row=row, column=last_col_idx)
        if isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.font = Font(color="0000FF", underline="single")

    # Sauvegarde finale
    wb.save("resultats_scenarios.xlsx")


    # Chargement du fichier
    wb = load_workbook("resultats_scenarios.xlsx")

    # Ajout d'une nouvelle feuille
    if "Details_Scenarios" in wb.sheetnames:
        ws = wb["Details_Scenarios"]
        wb.remove(ws)  # Nettoyage si déjà présent
    ws = wb.create_sheet("Details_Scenarios")

    # Style pour l'entête
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    row = 1
    for scenario in scenarios:
        meta = scenario[0]
        panneaux = scenario[1:]

        # Titre du scénario
        ws.cell(row=row, column=1, value=f"Scénario : {meta['libelleScenario']}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        # Infos globales
        ws.cell(row=row, column=1, value="Investissement (€)")
        ws.cell(row=row, column=2, value=meta["investissement"])
        ws.cell(row=row, column=3, value="Conso Maison (W)")
        ws.cell(row=row, column=4, value=meta["conso_maison_W"])
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = bold_font
        row += 2

        # En-têtes panneaux
        headers = ["Libellé", "Angle", "Aspect", "MaxPower", "TrackerType", "HorizonType"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = border
        row += 1

        # Infos par panneau
        for p in panneaux:
            ws.cell(row=row, column=1, value=p["libelle"])
            ws.cell(row=row, column=2, value=p["angle"])
            ws.cell(row=row, column=3, value=p["aspect"])
            ws.cell(row=row, column=4, value=p["MaxPower"])
            ws.cell(row=row, column=5, value=p["TrackerType"])
            ws.cell(row=row, column=6, value=p["horizonType"])
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
            row += 1

        row += 2  # Espace entre scénarios

 
#################

  
    ws_main = wb.worksheets[0]  # La première feuille (où sont listés les scénarios)

    # Création d'un dict pour accès rapide aux infos de scénario
    scenario_map = {}
    for scenario in scenarios:
        meta = scenario[0]
        panneaux = scenario[1:]
        texte =  f"Conso maison : {meta['conso_maison_W']} W"

        # Formatage des panneaux pour éviter trop de texte
        for i, p in enumerate(panneaux, 1):
            texte += f"\n• {p['libelle']}: {p['MaxPower']} kW, Angle {p['angle']}°, Aspect {p['aspect']}°, Track {p['TrackerType']}, Horizon {p['horizonType']} "
        
        scenario_map[meta['libelleScenario']] = texte

    # Appliquer les commentaires aux noms trouvés dans la colonne A
    for row in ws_main.iter_rows(min_row=2, min_col=1, max_col=1):
        cell = row[0]
        nom = str(cell.value)
        if nom in scenario_map:
            comment_text = scenario_map[nom]
            # Créer un commentaire
            comment = Comment(comment_text, "AutoGen")
            
            # Ajuster le commentaire (largeur et hauteur)
            comment.width = 300  # Largeur du commentaire (en pixels)
            comment.height = 150  # Hauteur du commentaire (en pixels)
            
            # Ajouter le commentaire à la cellule
            cell.comment = comment

  
    # Fonction pour convertir un numéro de colonne en lettre
    def column_index_to_letter(index):
        """Convertit un index de colonne (numérique) en lettre correspondante (ex: 1 -> 'A', 2 -> 'B')"""
        letter = ''
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter



    #wb = load_workbook("resultats_scenarios.xlsx")
    ws_main = wb.worksheets[0]  # La première feuille (où sont listés les scénarios)

    # Insertion de lignes en haut (avant la ligne 1)
    ws_main.insert_rows(1, 7)  # Insère 7 lignes au-dessus de la première ligne existante

    # Style pour les titres
    bold_font = Font(bold=True)

    # Ajout des données dans les nouvelles lignes
    ws_main.cell(row=1, column=1, value="Latitude")
    ws_main.cell(row=1, column=2, value=lat).font = bold_font

    ws_main.cell(row=2, column=1, value="Longitude")
    ws_main.cell(row=2, column=2, value=lon).font = bold_font

    ws_main.cell(row=3, column=1, value="Période")
    ws_main.cell(row=3, column=2, value=f"{startyear} - {endyear}").font = bold_font

    ws_main.cell(row=4, column=1, value="Prix kWh (€)")
    ws_main.cell(row=4, column=2, value=prix_kw).font = bold_font

    ws_main.cell(row=5, column=1, value="Date de départ cumul")
    ws_main.cell(row=5, column=2, value=date_departcumul.date()).font = bold_font

    ws_main.cell(row=6, column=1, value="Consommation chauffe-eau journalière (kWh)")
    ws_main.cell(row=6, column=2, value=conso_chauffeau_journalière).font = bold_font

    ws_main.cell(row=7, column=1, value="Consommation maison(W)")
    ws_main.cell(row=7, column=2, value=conso_maison_W).font = bold_font

#
    ws_main.cell(row=1, column=5, value="installation(€)")
    ws_main.cell(row=1, column=6, value=installation).font = bold_font

    ws_main.cell(row=2, column=5, value="fixationPV(€)")
    ws_main.cell(row=2, column=6, value=fixationPanneau).font = bold_font
        
    ws_main.cell(row=3, column=5, value="µ_onduleur1voie(€)")
    ws_main.cell(row=3, column=6, value=microonduleur1voie).font = bold_font

    ws_main.cell(row=4, column=5, value="µ_onduleur2voie(€)")
    ws_main.cell(row=4, column=6, value=microonduleur2voie).font = bold_font

    ws_main.cell(row=5, column=5, value="tracker(€)")
    ws_main.cell(row=5, column=6, value=tracker).font = bold_font

    ws_main.cell(row=6, column=5, value="Panneau(€)")
    ws_main.cell(row=6, column=6, value=PV).font = bold_font





    # Réappliquer le filtre à la nouvelle ligne 5 (au lieu de 1)
    last_column_letter = column_index_to_letter(ws_main.max_column)  # Convertir max_column en lettre
    ws_main.auto_filter.ref = f"A8:{last_column_letter}8"  # Définir la plage du filtre de la ligne 5

    # Sauvegarde du fichier
    wb.save(excelfilename)





def attendre_fermeture_fichier(fichier):
    while True:
        try:
            # On tente d'ouvrir en écriture exclusive
            with open(fichier, "a"):
                print(f"✅ Le fichier '{fichier}' est libre. On continue.")
                break
        except PermissionError:
            input(f"❌ Le fichier '{fichier}' est actuellement ouvert.\n👉 Merci de le fermer, puis appuyez sur [Entrée] pour réessayer...")
        time.sleep(1)



#----------------------------------------------
def calculate_solar_position(latitude, longitude, date):
    """Calcule la position solaire (azimut et élévation) pour une date et lieu donnés"""
    # Convertir en radians
    lat_rad = math.radians(latitude)
    
    # Jour de l'année
    day_of_year = date.timetuple().tm_yday
    
    # Déclinaison solaire
    delta = math.radians(23.45 * math.sin(math.radians((360.0/365.0) * (day_of_year - 81))))
    
    # Equation du temps (approximation)
    B = math.radians((360.0/365.0) * (day_of_year - 81))
    ET = 9.87 * math.sin(2*B) - 7.53 * math.cos(B) - 1.5 * math.sin(B)
    
    # Heure locale
    hour = date.hour + date.minute/60.0 + date.second/3600.0
    
    # Correction pour la longitude
    meridian = 15 * round(longitude/15)  # Méridien standard le plus proche
    LC = (meridian - longitude) / 15.0
    
    # Temps solaire
    solar_time = hour + ET/60.0 + LC
    
    # Angle horaire (en radians)
    hour_angle = math.radians(15.0 * (solar_time - 12.0))
    
    # Altitude (élévation)
    sin_altitude = (math.sin(lat_rad) * math.sin(delta) + 
                   math.cos(lat_rad) * math.cos(delta) * math.cos(hour_angle))
    altitude = math.degrees(math.asin(sin_altitude))
    
    # Azimut
    cos_azimuth = ((math.sin(delta) * math.cos(lat_rad) - 
                   math.cos(delta) * math.sin(lat_rad) * math.cos(hour_angle)) / 
                   math.cos(math.radians(altitude)))
    
    # Limiter cos_azimuth entre -1 et 1 pour éviter les erreurs numériques
    cos_azimuth = max(-1.0, min(1.0, cos_azimuth))
    
    azimuth = math.degrees(math.acos(cos_azimuth))
    
    # Ajuster l'azimut (0° au nord, sens horaire)
    if hour_angle > 0:
        azimuth = 360.0 - azimuth
    
    return azimuth, altitude

def get_sun_path_for_day(latitude, longitude, date):
    """Calcule la trajectoire du soleil pour une journée entière"""
    azimuths = []
    elevations = []
    
    # Calculer la position solaire toutes les 10 minutes
    for hour in range(24):
        for minute in range(0, 60, 10):
            current_time = date.replace(hour=hour, minute=minute, second=0)
            azimuth, elevation = calculate_solar_position(latitude, longitude, current_time)
            
            # Ne garder que les points où le soleil est au-dessus de l'horizon
            if elevation > 0:
                azimuths.append(azimuth)
                elevations.append(elevation)
    
    return azimuths, elevations

def create_solar_diagram(latitude, longitude, horizon_mask, save_path="diagramme_solaire.png"):
    """Crée et enregistre le diagramme solaire avec les trajectoires du soleil"""
    if len(horizon_mask) != 361:
        raise ValueError("Le masque d'horizon doit contenir 361 valeurs (azimut 0° à 360°).")
    
    # Configuration du graphique polaire
    fig, ax = plt.subplots(subplot_kw={'projection': 'polar'}, figsize=(10, 8))
    ax.set_theta_zero_location("N")
    ax.set_theta_direction(-1)
    
    # Tracer le masque d'horizon
    azimuths_deg = np.arange(0, 361)
    azimuths_rad = np.radians(azimuths_deg)
    elevation = 90 - np.array(horizon_mask)  # pour projection polaire
    ax.plot(azimuths_rad, elevation, label="Masque d'horizon", color="black", linewidth=2)
    
    # Dates pour chaque mois (21ème jour pour être proche des solstices/équinoxes)
    year = 2025  # Année actuelle
    dates = [datetime(year, month, 21) for month in range(1, 13)]
    
    # Couleurs pour chaque mois
    colors = plt.cm.hsv(np.linspace(0, 1, 12))
    
    # Tracer les trajectoires du soleil pour chaque mois
    for i, date in enumerate(dates):
        azimuths, elevations = get_sun_path_for_day(latitude, longitude, date)
        if azimuths:  # Vérifier qu'il y a des points à tracer
            # Pour l'affichage polaire
            radii = 90 - np.array(elevations)  
            theta = np.radians(azimuths)
            
            # Tracer la course du soleil
            line = ax.plot(theta, radii, color=colors[i], linewidth=1.5, 
                         label=date.strftime("%B"))
            
            # Marquer les heures (tous les 2 heures entre 6h et 18h)
            for hour in [6, 8, 10, 12, 14, 16, 18]:
                hour_date = date.replace(hour=hour, minute=0, second=0)
                azimuth, elevation = calculate_solar_position(latitude, longitude, hour_date)
                if elevation > 0:  # Si le soleil est visible
                    radius = 90 - elevation
                    ax.text(np.radians(azimuth), radius, f"{hour}", 
                           fontsize=8, ha='center', va='center', 
                           bbox=dict(facecolor='white', alpha=0.7, pad=1))

        # Hide the default tick labels
    # Remove all automatic ticks and labels
    ax.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)

    
    # Configuration des limites et étiquettes
    ax.set_rlim(0, 90)
    ax.set_rticks([15, 30, 45, 60, 75, 90])
    ax.set_rlabel_position(135)
     
    # Ajouter les étiquettes cardinales
    ax.set_xticklabels(['N', 'NE', 'E', 'SE', 'S', 'SO', 'O', 'NO'])
    ax.set_yticklabels(['75°', '60°', '45°', '30°', '15°', '0°'])  # Etiquettes d'élévation
    
    # Titre et légende
    plt.title(f"Diagramme solaire avec trajectoires mensuelles\nLatitude: {latitude}°, Longitude: {longitude}°", 
             y=1.08)
    plt.legend(loc='upper center', bbox_to_anchor=(0.5, -0.05), ncol=4)
    
    # Ajouter une grille
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    
    # Enregistrer l'image
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    return save_path



def calculate_sun_positions_for_year(latitude, longitude):
    """Calcule les positions solaires pour toute l'année, chaque heure entre le lever et le coucher du soleil"""
    # Année courante
    year = datetime.now().year
    
    # Structure pour stocker les résultats
    results = []
    
    # Pour chaque jour de l'année
    for month in range(1, 13):
        # On prend un jour au milieu du mois pour représenter le mois
        if month == 2 and year % 4 == 0 and (year % 100 != 0 or year % 400 == 0):
            days_in_month = 29
        elif month == 2:
            days_in_month = 28
        elif month in [4, 6, 9, 11]:
            days_in_month = 30
        else:
            days_in_month = 31
        
        # Pour le 21ème jour de chaque mois (proche des solstices/équinoxes)
        day = 21
        
        # Pour chaque heure de la journée
        for hour in range(24):
            for minute in [0, 30]:  # Toutes les 30 minutes
                date = datetime(year, month, day, hour, minute)
                azimuth, elevation = calculate_solar_position(latitude, longitude, date)
                
                # Ne garder que les positions où le soleil est au-dessus de l'horizon
                if elevation > 0:
                    results.append({
                        'date': date,
                        'month': month,
                        'hour': hour + minute/60.0,
                        'azimuth': azimuth,
                        'elevation': elevation
                    })
    
    return pd.DataFrame(results)

def check_horizon_blocking(azimuth, elevation, horizon_mask):
    """Vérifie si le soleil est bloqué par l'horizon à cette position"""
    # Trouver l'élévation de l'horizon à cette position d'azimut
    azimuth_idx = int(round(azimuth))
    if azimuth_idx >= len(horizon_mask):
        azimuth_idx = len(horizon_mask) - 1
    horizon_elevation = horizon_mask[azimuth_idx]
    
    # Le soleil est visible si son élévation est supérieure à celle de l'horizon
    return elevation > horizon_elevation


def get_smooth_sun_positions_for_day(latitude, longitude, month, day=21, interval_minutes=15):
 
    
    # Créer une journée typique à intervalle régulier
    date = datetime(2022, month, day)
    times = pd.date_range(start=date, periods=int(24*60/interval_minutes), freq=f"{interval_minutes}min", tz='UTC')
    
    # Calculer les positions solaires
    solpos = solarposition.get_solarposition(times, latitude, longitude)
    
    # Ajouter les colonnes utiles
    df = pd.DataFrame({
        'time': times,
        'azimuth': solpos['azimuth'].values,
        'elevation': solpos['apparent_elevation'].values,
        'month': month
    })
    return df



def create_cartesian_solar_diagram(i, latitude, longitude, horizon_mask, save_path="diagramme_solaire_cartesien.png"):
    """Crée un diagramme solaire cartésien et l'enregistre"""
    save_path = str(i) + save_path

    # Calculer les positions solaires pour toute l'année
    sun_positions = calculate_sun_positions_for_year(latitude, longitude)

    # Configuration du graphique - plus large que haut
    fig = plt.figure(figsize=(10, 8))

    # Single plot
    ax1 = fig.add_subplot(111)

    # Sélectionner quelques jours représentatifs (par exemple, le 21 de chaque mois)
    selected_months = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    colors = plt.cm.hsv(np.linspace(0, 1, len(selected_months)))

    # Pour chaque mois, tracer la trajectoire du soleil
    for i, month in enumerate(selected_months):
#        month_data = sun_positions[sun_positions['month'] == month]
        month_data = get_smooth_sun_positions_for_day(latitude, longitude, month)

        if not month_data.empty:
            # Déterminer si les points sont bloqués par l'horizon
            visible_points = []
            for _, row in month_data.iterrows():
                visible = check_horizon_blocking(row['azimuth'], row['elevation'], horizon_mask)
                visible_points.append(visible)

            visible_points = np.array(visible_points)

            # Tracer les points visibles
            visible_data = month_data[visible_points]
            ax1.plot(visible_data['azimuth'], visible_data['elevation'], 'o-',
                     color=colors[i], label=f"Mois {month}", markersize=2)

         
            # Ajouter des annotations d'heure à certains points, sans doublons
            heures_affichees = set()
            for _, row in visible_data.iterrows():
                heure = pd.to_datetime(row['time']).hour
                if heure in [6,8,10, 12, 14, 16, 18, 20] and heure not in heures_affichees:
                    ax1.text(row['azimuth'], row['elevation'] + 1, f"{heure}h", 
                             fontsize=6, ha='center', color=colors[i])
                    heures_affichees.add(heure)



            # Tracer les points bloqués
            blocked_data = month_data[~visible_points]
            if not blocked_data.empty:
                ax1.plot(blocked_data['azimuth'], blocked_data['elevation'], 'o',
                         color=colors[i], alpha=0.3, markersize=2)

    # Tracer le masque d'horizon
    azimuths = np.arange(0, 361)
    ax1.plot(azimuths, horizon_mask, 'k-', linewidth=2, label="Horizon")

    # Configuration des axes et légendes
    ax1.set_xlim(40, 320)  # Zoom horizontal entre 40 et 320
    ax1.set_ylim(0, max(90, max(sun_positions['elevation']) * 1.1))
    ax1.set_xlabel('Azimut (degrés)')
    ax1.set_ylabel('Élévation (degrés)')
    ax1.set_title('Trajectoires solaires par mois')

    # Ticks personnalisés
    azimuth_labels = range(40, 321, 20)
    ax1.set_xticks(azimuth_labels)
    ax1.set_xticklabels([f"{az}°" for az in azimuth_labels])

    ax1.set_yticks(np.arange(0, 91, 10))
    ax1.grid(True, alpha=0.3)
    ax1.xaxis.grid(True, which='major', linestyle='-', linewidth=0.8, alpha=0.5)

    ax1.legend(loc='upper right', ncol=2)

    # Adapter la forme du graphe à l'espace disponible
    ax1.set_aspect('auto')

    # Ajustement serré pour coller les bords
    plt.subplots_adjust(left=0.1, right=0.9, bottom=0.1, top=0.9)

    # Enregistrement sans bbox_inches
    plt.savefig(save_path, dpi=600)

    return save_path




def add_diagram_to_excel(excel_file,i,  image_path ,latitude , longitude ):
    """Ajoute le diagramme solaire à un nouvel onglet dans le fichier Excel"""
    # Charger le fichier Excel existant
    workbook = load_workbook(excel_file)
    
    # Créer un nouvel onglet pour le diagramme
    sheet_name = "Diagramme Solaire_"+str(i)
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)
    
    # Ajouter un titre
    worksheet['A1'] = "Diagramme Solaire - Course du soleil mois par mois"
    worksheet['A2'] = f"Coordonnées: Latitude {latitude}°, Longitude {longitude}°"
    
    # Insérer l'image
    img = Image(image_path)
    
    # Ajuster la taille de l'image si nécessaire (optionnel)
    scale_factor = 0.4  # Réduire à 75% de la taille originale
    img.width = img.width * scale_factor
    img.height = img.height * scale_factor
    
    # Positionner l'image sous le titre
    worksheet.add_image(img, 'A4')
    
    # Enregistrer le fichier Excel modifié
    workbook.save(excel_file)
    
    return True
        
#-------------------------------------------------------
#main



# Exemple d'utilisation
fichier = "resultats_scenarios.xlsx"
if os.path.exists(fichier):
    attendre_fermeture_fichier(fichier)


# Traitement de chaque scénario
for scenario in scenarios:
    meta = scenario[0]
    configs = scenario[1:]
   
    atleastone=doaall    #mettre 1 pour clean

    print(f"\n▶️ Scénario: {meta['libelleScenario']}")
    for config in configs:
        print(f"  - Panneau: angle={config['angle']}°, aspect={config['aspect']}°, power={config['MaxPower']} kW")

    df_cumul = pd.DataFrame()
    dfs_scenarios = {}

    nompourhash=""
    for i, params in enumerate(configs, 1):
        url, filename = build_pvgis_url(i,lat, lon, **params, startyear=startyear, endyear=endyear)
        csvfile= "pvgisdata"+filename+".csv"
        nom_finalbrut = f"energie_brut"+filename+".xlsx"

        nompourhash=nompourhash+nom_finalbrut


       
        updated = telecharger_csv(url, csvfile)


 
  
        if not os.path.exists(nom_finalbrut):
             print("not exist ?",nom_finalbrut)
             updated=1
             
        if updated == 1    :
            atleastone=1
            # 🔄 Fichier mis à jour → retraitement du CSV
            brut = traitement_csv(csvfile, params["MaxPower"])
            brut["datetime"] = pd.to_datetime(brut["datetime"])
            brut = brut[["datetime", "p"]]
            brut.to_excel(nom_finalbrut,index=False)

        else:
            # 📂 Fichier déjà présent → relire Excel existant
            print("relecture",nom_finalbrut)
            brut = pd.read_excel(nom_finalbrut)
  
           # brut = brut[["datetime", "p"]]
             
        brut.set_index('datetime', inplace=True)
        dfs_scenarios[f"scenario_{i}"] = brut.copy()
        # ⚙️ Ajouter les puissances au cumul
        df_cumul = brut if df_cumul.empty else df_cumul.add(brut, fill_value=0)

    # Ajouter chaque scénario en colonne individuelle
    for name, df in dfs_scenarios.items():
        df_cumul[f"{name}"] = df["p"]

    # Export du fichier cumulatif brut
    thach= hashlib.sha256(nompourhash.encode()).hexdigest()[:8]    
    nom_scenario = f"energie_scenarios_{meta['libelleScenario']}_{thach}.xlsx"
    nom_final = f"energie_cumulee_scenarios_{meta['libelleScenario']}_{thach}.xlsx"

    if atleastone  or  not os.path.exists(nom_scenario) :
        print("regenere cumul pour ce scenario:",nom_scenario)
        df_cumul.to_excel(nom_scenario)

    # Traitement final
        df_cumulfin = traitement_cumul(nom_scenario, meta['conso_maison_W'])
        df_cumulfin.to_excel(nom_final)
        ajouter_graphique_excel(nom_final, df_cumulfin)
    
    else:
        df_cumulfin = pd.read_excel(nom_final)

   
    last_row = df_cumulfin[["cumul_besoin (Kwh)", "energie_perdue"	,"manque",	"energie_recuperee",  "cumul_gain (€)", "cumul_perte (€)" ,  "cumul_manque (€)" ]].iloc[-1]

    besoin=last_row["cumul_besoin (Kwh)"]
    gain = last_row["cumul_gain (€)"]
    perte = last_row["cumul_perte (€)"]
    manque =last_row["cumul_manque (€)"]
    
    investissement = meta["investissement"]

    ratio = perte / gain if gain else None
    gain_sur_invest = gain / investissement if investissement else None

    resultats.append({
        "scenario": meta["libelleScenario"],
        "investissement": investissement,
         "BesoinTotal(Kwh)": besoin,
        "gain(Kwh)": gain / prix_kw ,
        "perte(Kwh)":perte / prix_kw,
        "manque(Kwh)": manque / prix_kw,
        
        "BesoinTotal(€)": besoin * prix_kw ,
        "gain(€)": gain,
        "perte(€)": perte,
        "manque(€)": manque,
 
        "ratio_perte_gain": ratio,
        "gain_sur_invest": gain_sur_invest,
      "Depense": investissement+ manque,
      "Gain_Sur_depense": gain/ (investissement+ manque),


        
        "lien_excel": f'=HYPERLINK("{nom_final}", "Ouvrir Excel")'
    })


# Résumé final
print("\n📊 Résumé des scénarios :\n")
for res in resultats:
    print(f"🔹 {res['scenario']}")
    print(f"    Investissement     : {res['investissement']} €")
    print(f"    Gain cumulé        : {res['gain(€)']:.2f} €")
    print(f"    Perte cumulée      : {res['perte(€)']:.2f} €")
    print(f"    manque cumulée      : {res['manque(€)']:.2f} €")
    print(f"    Ratio perte/gain   : {res['ratio_perte_gain']:.2%}" if res['ratio_perte_gain'] is not None else "    ⚠️ Ratio perte/gain : gain = 0")
    print(f"    Gain / Invest      : {res['gain_sur_invest']:.2%}" if res['gain_sur_invest'] is not None else "    ⚠️ Impossible de calculer le rendement")
    print("")

# Export final des résultats
df_resultats = pd.DataFrame(resultats)

# 🔽 Appliquer la règle d’arrondi

def format_number(val):
        try:
            if isinstance(val, (int, float)):
                if abs(val) < 100:
                    return round(val,2)
                else:
                    return round(val)
            return val
        except:
            return val

df_resultats = df_resultats.applymap(format_number)

# Sauvegarde initiale avec pandas
saveto="resultats_scenarios.xlsx"
df_resultats.to_excel(saveto, index=False)
format_excel(saveto)

print("✅ Résumé global exporté dans 'resultats_scenarios.xlsx'")



for i in range(1, len(datahorizonType)):
    tableau = [float(x) for x in datahorizonType[i].split(',')]
    # Dupliquer le premier élément
    premier_element = tableau[0]
    tableau.insert(0, premier_element)  # Insérer une copie du premier élément au début
    # Créer et enregistrer le diagramme solaire
    diagram_path = create_solar_diagram(lat, lon, tableau)
    # Ajouter le diagramme à un nouvel onglet dans le fichier Excel
    add_diagram_to_excel(saveto, i, diagram_path, lat, lon)
    print(f"Le diagramme solaire pour l'horizon {i} a été ajouté avec succès au fichier Excel '{saveto}'.")
    
    diagram_path = create_cartesian_solar_diagram(i,lat, lon, tableau)
    add_diagram_to_excel(saveto, i+100, diagram_path, lat, lon)

    print(f"Diagramme enregistré sous: {diagram_path}")

    
if os.path.exists(fichier):
    print(f"📂 Ouverture de {fichier}...")
    os.startfile(fichier)
    

