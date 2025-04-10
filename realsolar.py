import importlib
import subprocess
import sys

# Liste des bibliothèques nécessaires
required_libs = [
    "pandas",
    "openpyxl",
    "os",
    "calendar",
    "urllib.parse",
    "requests",
    "urllib3",
    "datetime",
    "hashlib",
    "time",
    "matplotlib",
    "math"
]

# Fonction pour vérifier et installer les bibliothèques manquantes
def check_and_install(package_name):
    try:
        importlib.import_module(package_name.split('.')[0])
    except ImportError:
        print(f"Installation de {package_name}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])

# Vérifie toutes les dépendances
for lib in required_libs:
    check_and_install(lib)


import pandas as pd
import os

import calendar
import urllib.parse
import requests
import urllib3
from datetime import datetime
import hashlib
import time
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.collections import LineCollection
from datetime import  timedelta
import math

from openpyxl.drawing.image import Image





urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

############################################################################################"
#USER DATA

doaall=0;   # mettre a un pour ignorer les fichiers existants et tout refaire

# Variables globales
resultats = []
lat, lon = 45.762, 4.698
startyear, endyear = 2018, 2023

prix_kw = 0.2016  # €/kWh
date_departcumul = pd.Timestamp("2018-01-01")
conso_chauffeau_journalière = 2.2  # kWh

conso_maison_W= 100
# Définition des masks d'horizon

#attention si les data de l'horizon change il faut efface les fichiers intermediare ( faire doaall=1)  pour refaire la demande a pvgis
## definir un profile par lieu de panneau ( hauteur des masques de 0 (nord) a 360 degree )



datahorizonType=[
    "" ,
    #1:balcon
     "4.7266,4.7253,4.7241,4.7228,4.7215,4.7203,4.719,4.7178,4.7165,4.7153,4.714,4.7127,4.7115,4.7102,4.709,4.7077,4.7065,4.7052,\
4.704,4.7027,4.7014,4.7002,4.6989,4.6977,4.6964,4.6952,4.6939,4.6926,4.6914,4.6901,4.6889,4.6876,4.6864,4.6851,4.6839,\
4.6826,4.6813,4.6801,4.6788,4.6776,4.6763,4.6751,4.6738,4.6725,4.6713,4.67,4.6688,4.6675,4.6663,4.665,4.6638,4.6625,\
4.6612,4.66,4.6587,4.6575,4.6562,4.655,4.6537,4.6524,4.6512,4.6499,4.6487,4.6474,4.6462,4.6449,4.6437,4.6424,4.6411,\
4.6399,4.6386,4.6374,4.6364,4.6364,4.596,4.4343,3.9899,5.7273,5.9697,7.7072,10.9577,16.7717,18.4817,21.5892,24.2871,\
25.4848,25.5657,25.3232,24.8788,24.5152,24.0707,23.6667,22.9838,22.4949,22.1981,21.9871,21.202,20.555,21.3924,22.7156,\
25.8357,25.8523,25.2424,24.9454,25.2484,26.6162,27.7887,28.5556,27.9711,21.0061,15.8723,14.6566,15.4791,16.3365,17.5901,\
19.5455,19.9495,20.413,21.202,21.5623,22.1369,21.3086,20.6768,20.6286,21.2828,20.6768,19.0625,17.5778,17,16.3535,15.5306,\
14.956,13.8964,10.9726,9.404,8.2144,6.6294,3.6726,3.1414,4.3282,7.8196,9.382,10.6582,11.1521,11.4242,10.8299,10.4114,\
12.506,14.7329,16.095,20.1776,22.214,23.5455,24.2522,25.0289,25.6061,26.2994,27.4148,28.3131,29.2355,29.2828,29.2828,\
29.5253,29.8081,29.7273,29.6869,29.6465,29.6307,30.6903,32.1538,33.8016,35.303,35.6152,35.907,35.9091,35.8528,34.47,\
32.21,28.4268,25.837,24.3125,22.4064,20.7999,18.6756,11.3115,4.2378,1.6402,1.6869,1.7677,1.9697,1.7677,1.7677,1.8154,\
1.8485,1.9697,2.4485,2.4112,2.2525,2.4351,2.7374,2.7374,2.6162,2.899,3.0202,3.6263,3.8251,3.3434,3.2222,3.2524,3.3838,\
3.1414,3.1414,3.652,3.5051,3.4242,3.3838,3.4646,4.4243,5.9697,4.4747,4.2154,3.9899,3.8687,4.1515,4.1515,3.9495,3.8283,\
3.8687,4.0303,4.1111,4.1111,4.3939,4.4747,4.5152,4.5152,4.5132,4.4747,4.7172,4.8384,5.202,5.202,5.206,5.4444,5.404,5.404,\
5.6461,5.7677,5.7677,5.6061,5.4848,5.5657,5.7677,5.4848,5.5253,5.8081,6.0593,7.4242,7.101,7.5721,7.267,8.2995,6.899,6.4141,\
6.4949,8.6454,8.4567,6.7778,7.7475,7.7521,8.5288,9.2828,8.7664,8.1917,7.7787,6.5758,9.0272,9.5253,8.9596,8.1653,8.7804,\
7.9177,9.1212,9.6653,9.2424,8.4069,7.6706,7.5051,7.5051,8.1919,8.2323,8.1919,8.3939,7.7475,7.8559,8.1919,8.1164,8.5746,\
6.899,5.9537,5.2012,5.1616,5.1616,4.8384,4.7969,4.7957,4.7944,4.7932,4.7919,4.7906,4.7894,4.7881,4.7869,4.7856,4.7844,\
4.7831,4.7818,4.7806,4.7793,4.7781,4.7768,4.7756,4.7743,4.7731,4.7718,4.7705,4.7693,4.768,4.7668,4.7655,4.7643,4.763,\
4.7617,4.7605,4.7592,4.758,4.7567,4.7555,4.7542,4.753,4.7517,4.7504,4.7492,4.7479,4.7467,4.7454,4.7442,4.7429,4.7416,\
4.7404,4.7391,4.7379,4.7366,4.7354,4.7341,4.7328,4.7316,4.7303,4.7291,4.7278,4.7266",

    #2:portail
    "14.3116,14.3401,14.3685,14.3970,14.4255,14.4540,14.4825,14.5109,14.5394,14.5679,14.5964,14.6248,14.6533,14.6818,14.7103,\
14.7388,14.7672,14.7957,14.8242,14.8527,14.8811,14.9096,14.9381,14.9666,14.9951,15.0235,15.0520,15.0805,15.1090,15.1374,15.1659,\
15.1944,15.2229,15.2514,15.2798,15.3083,15.3368,15.3653,15.3937,15.4222,15.4507,15.4792,15.5077,15.5361,15.5646,15.5931,15.6216,\
15.6500,15.6785,15.7070,15.7420,15.7840,15.8276,15.8154,15.7750,15.7735,15.8049,15.8614,15.8364,15.8364,15.7735,15.7944,15.8364,\
15.8783,15.9098,15.9518,15.9518,16.0615,16.0671,16.0042,15.9937,15.9622,15.9518,15.9832,16.0042,16.0776,16.1720,16.1720,16.1615,\
16.1196,16.0986,16.0461,15.9518,15.8888,15.7723,15.7315,15.7315,15.6829,15.4274,15.2213,13.5501,13.5081,13.0047,13.0113,13.6759,\
13.6759,13.2962,12.8729,13.6445,13.7416,12.7635,13.4662,13.6472,13.4581,12.7320,13.3740,13.2953,12.3335,13.1233,13.2040,13.1411,\
14.5988,14.8017,15.0273,15.2424,15.4784,15.5008,15.4076,15.4064,15.2920,15.0288,15.1442,14.8670,14.9030,14.5884,14.6408,14.3075,\
14.0010,14.0325,14.0632,14.6670,14.7352,14.7615,14.2106,13.9850,13.8962,13.9381,13.9270,13.9067,13.8332,13.7747,13.6121,13.9024,\
14.4134,15.2200,15.3749,15.3264,14.8881,14.0645,13.2829,10.7218,11.0540,11.1451,11.2428,11.6833,11.6938,11.7462,11.6833,11.7147,\
11.9795,12.3020,12.4306,12.9628,12.8235,12.7448,12.4489,12.3545,12.1867,12.1657,12.4418,12.6359,12.6272,12.5118,12.5890,13.0056,\
13.2918,13.5174,13.7430,13.9790,14.1793,14.4512,14.6872,14.8820,15.0393,15.1547,15.2596,15.3432,15.3540,15.4693,15.5637,15.6686,\
15.7420,15.8261,15.8623,15.8053,15.8154,16.0583,16.3188,16.3503,14.6888,14.7028,14.7834,15.0940,15.5516,15.7944,15.2673,14.9030,\
14.8401,14.9030,14.9135,14.9135,14.9764,15.0184,15.0603,15.0393,15.0079,15.0184,13.2104,11.3086,11.2847,11.3057,11.4711,11.8196,\
11.8406,11.8616,11.6224,11.3686,11.2447,11.1450,11.0243,11.1136,11.2113,11.1379,11.0855,11.2821,11.5050,11.4252,11.3791,11.0960,\
11.4622,11.6518,11.3791,11.2743,11.2533,11.4049,11.5679,12.0460,12.3651,12.2128,13.1620,13.0676,12.7320,12.3915,12.4230,11.3734,\
11.3582,11.3784,11.4069,11.4354,11.4638,11.4923,11.5208,11.5493,11.5777,11.6062,11.6347,11.6632,11.6917,11.7201,11.7486,11.7771,\
11.8056,11.8340,11.8625,11.8910,11.9195,11.9480,11.9764,12.0049,12.0334,12.0619,12.0903,12.1188,12.1473,12.1758,12.2043,12.2327,\
12.2612,12.2897,12.3182,12.3466,12.3751,12.4036,12.4321,12.4606,12.4890,12.5175,12.5460,12.5745,12.6029,12.6314,12.6599,12.6884,\
12.7168,12.7453,12.7738,12.8023,12.8308,12.8592,12.8877,12.9162,12.9447,12.9731,13.0016,13.0301,13.0586,13.0871,13.1155,13.1440,\
13.1725,13.2010,13.2294,13.2579,13.2864,13.3149,13.3434,13.3718,13.4003,13.4288,13.4573,13.4857,13.5142,13.5427,13.5712,13.5997,\
13.6281,13.6566,13.6851,13.7136,13.7420,13.7705,13.7990,13.8275,13.8560,13.8844,13.9129,13.9414,13.9699,13.9983,14.0268,14.0553,\
14.0838,14.1123,14.1407,14.1692,14.1977,14.2262,14.2546,14.2831,14.3116"
]

# Définition des scénarios de pose des panneaux

#note:
 # pas d'espace dans les libellé
 # tracking    0= fixe   2 2 axes    3 axe vertical qui tourne      5 axe incliné qui tourne
 # horizonType 0 = use default location   autre indice dans la table des profils d'horizon
 
scenarios = [

            [     
        {"libelleScenario": "x1Tracker2axes_balcon", "investissement": 120+80+45+100 ,"conso_maison_W":conso_maison_W },
        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 2 , "horizonType": 1 , "libelle": "tracker"},
    ],

        
     [  
        {"libelleScenario": "1xToitureEst_1xBalcon", "investissement": 120+80+80+190,"conso_maison_W":conso_maison_W},
         {"angle": 23, "aspect": -(180-72) , "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "est"},
         {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1 , "libelle": "balcon_72"},
 
    ],
     
    [  
        {"libelleScenario": "2xToitureOuest_sansmask", "investissement": 120+80+80+190,"conso_maison_W":conso_maison_W},
        {"angle": 23, "aspect": 72, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "ouest"},
        {"angle": 23, "aspect": 72, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "ouest"},
    ],

    [  
        {"libelleScenario": "1xToitureEst_1xToitureouest_sansmask", "investissement": 120+80+80+190,"conso_maison_W":conso_maison_W},
        {"angle": 23, "aspect": 72, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "ouest"},
        {"angle": 23, "aspect": -(180-72) , "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "est"},
    ],



     [  
        {"libelleScenario": "1xBalcon_1xBalconOuest", "investissement": 120+80+80+190,"conso_maison_W":conso_maison_W},
         {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1 , "libelle": "balcon_72"},
         {"angle": 72, "aspect": 72, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "ouest72"},
 
    ],
 

                        [     
        {"libelleScenario": "x1Tracker2axes_portail", "investissement": 120+80+45+100 ,"conso_maison_W":conso_maison_W },
        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 2 , "horizonType": 2 , "libelle": "tracker"},
    ],

                        [     
        {"libelleScenario": "x2Tracker_2axes_Portail", "investissement": 120+80+45+100 + 180+45+100  ,"conso_maison_W":conso_maison_W },
        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 2 , "horizonType": 2 , "libelle": "tracker"},
        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 2 , "horizonType": 2 , "libelle": "tracker"},
    ],
            
            
    [         
        {"libelleScenario": "1xPortail", "investissement": 120+ 80+126 ,"conso_maison_W":conso_maison_W},
        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2 , "libelle": "optimal"},
    ],

    [  
        {"libelleScenario": "1xBalcon", "investissement": 120+80+126 ,"conso_maison_W":conso_maison_W },
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1 , "libelle": "balcon_72"},
    ],




    [  
        {"libelleScenario": "2xBalcon", "investissement": 120+80+80+190,"conso_maison_W":conso_maison_W},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
    ],
    [  
        {"libelleScenario": "2xGarageOuest_sansmask", "investissement": 120+80+80+190,"conso_maison_W":conso_maison_W},
        {"angle": 23, "aspect": 47, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "garage_ouest"},
        {"angle": 23, "aspect": 47, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "garage_ouest"},
    ],
    [  # mix optimal + balcon 
        {"libelleScenario": "1xPortail_1xBalcon", "investissement": 120+80+80+126+126,"conso_maison_W":conso_maison_W},
        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimal"},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
    ],
    [  # optimal x2  
        {"libelleScenario": "2xPortail", "investissement": 120+80+80+190,"conso_maison_W":conso_maison_W},
        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimal"},
        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimal"},
    ],

##        [  # optimal x3
##        {"libelleScenario": "3x_optimal_sans_mask", "investissement": 120+80+80+80+190+125,"conso_maison_W":conso_maison_W},
##        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "optimal"},
##        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "optimal"},
##        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 0, "libelle": "optimal"},
##    ],

        [  #  mix 
        {"libelleScenario": "1xPortail_2xBalcon", "investissement": 120+80+80+190 +80+125,"conso_maison_W":conso_maison_W},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
        {"angle": 72, "aspect": -18, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 1, "libelle": "balcon_72"},
        {"angle": 37, "aspect": -2, "MaxPower": 0.5, "TrackerType": 0 , "horizonType": 2, "libelle": "optimal"},
 
    ],
    
##         [  # optimal pas d'ombre
##        {"libelleScenario": "4xOptimal_sans_mask", "investissement": 120+80+80+80+80+ 280,"conso_maison_W":conso_maison_W},
##        {"angle": 37, "aspect": -2, "MaxPower": 0.52, "TrackerType": 0 , "horizonType": 0, "libelle": "optimal"},
##        {"angle": 37, "aspect": -2, "MaxPower": 0.52, "TrackerType": 0 , "horizonType": 0, "libelle": "optimal"},
##        {"angle": 37, "aspect": -2, "MaxPower": 0.52, "TrackerType": 0 , "horizonType": 0, "libelle": "optimal"},
##         {"angle": 37, "aspect": -2, "MaxPower": 0.52, "TrackerType": 0 , "horizonType": 0, "libelle": "optimal"},
##       
##    ],   
]

#END USER DATA
############################################################################################"





def build_pvgis_url(i,lat, lon, angle, aspect,libelle, startyear, endyear, TrackerType, horizonType ,  MaxPower=1.00, tech='crystSi', db='PVGIS-SARAH3'):
    base_url = "https://re.jrc.ec.europa.eu/api/v5_3/seriescalc"
    params = {
        "lat": lat,
        "lon": lon,
        "raddatabase": db,
        "outputformat": "csv",
        "usehorizon": 1,
        "angle": angle,
        "aspect": aspect,
        "startyear": startyear,
        "endyear": endyear,
        "mountingplace": "free",
        "trackingtype": TrackerType,              #0= fixe   2 2 axes    3 axe vertical qui tourne      5 axe incliné qui tourne 
        "pvcalculation": 1,
        "pvtechchoice": tech,
        "peakpower": MaxPower,
        "loss": 14,
        "select_database_hourly":"PVGIS-SARAH3",

        
    }
    if horizonType != 0 :
        params["userhorizon"] = datahorizonType[horizonType]

        
    url = base_url + "?" + urllib.parse.urlencode(params)

    filename = f"idata{i}_{angle}deg_{aspect}deg_{startyear}_{endyear}_PMax{MaxPower:.2f}_tracking{TrackerType}_hor{horizonType}.csv"
    return url, filename


def telecharger_csv(url, filename):
    if not os.path.exists(filename)  or doaall==1:
        print(f"Téléchargement depuis : {url}")
        response = requests.get(url, verify=False)
        if response.ok:
            with open(filename, 'wb') as f:
                f.write(response.content)
            print(f"✅ Fichier enregistré sous : {filename}")
            return 1
        else:
            print(f"❌ Erreur {response.status_code}")
            return 1
    else:
        print(f"✔️ Fichier déjà existant : {filename}")
        return 0


def traitement_csv(filename, maxpower):
    with open(filename, 'r', encoding='utf-8') as f:
        lignes = f.readlines()
    for i, ligne in enumerate(lignes):
        if ligne.strip().startswith("time,P"):
            header_idx = i
            break

    df = pd.read_csv(filename, skiprows=header_idx)
    df.columns = df.columns.str.strip().str.lower()
    df[['date', 'time']] = df['time'].str.split(":", expand=True)
    df = df[df['date'].astype(str).str.match(r'^\d{8}$')]
    df['date'] = pd.to_datetime(df['date'], format='%Y%m%d')
    df['time'] = df['time'].astype(str).str.zfill(4)
    df['hour'] = df['time'].str[:-2].astype(int)
    df['minute'] = df['time'].str[-2:].astype(int)
    df['datetime'] = df['date'] + pd.to_timedelta(df['hour'], unit='h') + pd.to_timedelta(df['minute'], unit='m')
    df['p'] = pd.to_numeric(df['p'], errors='coerce')

    return df






def traitement_cumul(filename,conso_maison_W):

    df = pd.read_excel(filename)
    conso_maison_kW = conso_maison_W / 1000
        
    df['date'] = pd.to_datetime(df['datetime']).dt.date
   
    df['p_dispo'] = 0.0
    df['p_maison'] = 0.0
    df['p_chauffeau'] = 0.0
    df['p_perdue'] = 0.0

    df['p_besoin'] = 0.0
      
    for date_jour, df_jour in df.groupby(df['datetime'].dt.date):
        energie_chauffeau_cumul = 0.0
        for i, row in df_jour.iterrows():
            p_dispo = row['p']
            df.at[i, 'p_dispo']=p_dispo
            if p_dispo >= conso_maison_W:
                df.at[i, 'p_maison'] = conso_maison_W
                p_rest = p_dispo - conso_maison_W
            else:
                df.at[i, 'p_maison'] = p_dispo
                p_rest = 0

            if energie_chauffeau_cumul < conso_chauffeau_journalière:
                reste = (conso_chauffeau_journalière - energie_chauffeau_cumul) * 1000
                chauffe_eau = min(p_rest, reste)
                df.at[i, 'p_chauffeau'] = chauffe_eau
                energie_chauffeau_cumul += chauffe_eau / 1000
                p_rest -= chauffe_eau

            df.at[i, 'p_perdue'] = p_rest
           
            df.at[i,  'p_besoin'] =    conso_chauffeau_journalière*1000 /24  + conso_maison_W  


    df['date'] = pd.to_datetime(df['datetime'].dt.date)

    # Étape 1 : Identifier dynamiquement les colonnes 'scenario_*'
    scenario_cols = [col for col in df.columns if col.startswith('scenario_')]

    # Étape 2 : Ajouter les autres colonnes fixes
    other_cols = [ 'p_besoin', 'p_dispo', 'p_maison', 'p_chauffeau', 'p_perdue']

    # Construction du dictionnaire d'aggregation
    agg_dict = {col: 'sum' for col in scenario_cols + other_cols}

    # Étape 3 : Groupby + agg
    daily = df.groupby('date').agg(agg_dict) / 1000
          
 
 
    daily['manque'] = (
            (conso_chauffeau_journalière + conso_maison_kW * 24 ) -   #  * df.groupby('date').size() - 1.5  ) -         # on envele 2 heure
            (daily['p_chauffeau'] + daily['p_maison'])
        ).clip(lower=0)



    # Ajustement: si 'energie_perdue' > 0, alors 'manque' = 0
    daily.loc[daily['p_perdue'] > 0, 'manque'] = 0

    daily = daily.rename(columns={
      
        'p_besoin': 'besoinKwh',
        'p_maison': 'energie_maison',
        'p_chauffeau': 'energie_chauffeau',
        'p_perdue': 'energie_perdue'
    })
    daily['energie_recuperee'] = daily['energie_maison'] + daily['energie_chauffeau']

    daily['gain (€)'] = daily['energie_recuperee'] * prix_kw
    daily['perte (€)'] = daily['energie_perdue'] * prix_kw
    daily['manque (€)'] = daily['manque'] * prix_kw

    daily['cumul_gain (€)'] = daily[daily.index >= date_departcumul]['gain (€)'].cumsum()
    daily['cumul_perte (€)'] = daily[daily.index >= date_departcumul]['perte (€)'].cumsum()

    daily['cumul_manque (€)'] = daily[daily.index >= date_departcumul]['manque (€)'].cumsum()

    daily['cumul_besoin (Kwh)'] = daily[daily.index >= date_departcumul]['besoinKwh'].cumsum()


    daily.loc[daily.index < date_departcumul, ['cumul_gain (€)', 'cumul_perte (€)']] = float('nan')

    
    return daily




def ajouter_graphique_excel(fichier_excel, df_resultat):
    wb = load_workbook(fichier_excel)
    ws = wb.active

    # Recherche de l’index (colonne) de 'energie_maison'
    header = [cell.value for cell in ws[1]]
    try:
        col_energie_maison = header.index("energie_maison") + 1  # Excel est 1-based
    except ValueError:
        print("⚠️ Colonne 'energie_maison' non trouvée dans le fichier Excel.")
        return

    # Conversion des dates pour grouper par mois
    df_resultat["date"] = pd.to_datetime(df_resultat.index)
    df_resultat["mois"] = df_resultat["date"].dt.to_period("M")

    start_row = 2
    row_offset = 10
    chart_height = 5
    chart_width = 20

    for i, mois in enumerate(df_resultat["mois"].unique()):
        df_mois = df_resultat[df_resultat["mois"] == mois]
        excel_start = df_resultat.index.get_loc(df_mois.index[0]) + 2
        excel_end = excel_start + len(df_mois) - 1

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = f"{calendar.month_name[mois.month]} {mois.year}"
        chart.y_axis.title = "Énergie (kWh)"
        chart.x_axis.title = "Jour"

        # Utilise la colonne 'energie_maison' et celle juste après (si tu veux plusieurs)
        data = Reference(ws,
                         min_col=col_energie_maison,
                         max_col=col_energie_maison + 3,  # Ajuste selon besoin
                         min_row=excel_start - 1,
                         max_row=excel_end)

        # Catégories (date/jour)
        cats = Reference(ws, min_col=1, max_col=1, min_row=excel_start, max_row=excel_end)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = chart_height
        chart.width = chart_width

        ws.add_chart(chart, f"Q{start_row + i * row_offset}")

    wb.save(fichier_excel)




def format_excel(excelfilename):
       
    # Ouvre le fichier pour modifications avec openpyxl
    wb = load_workbook(excelfilename)
    ws = wb.active

    # 1. Ajoute une ligne de filtrage (auto-filter)
    ws.auto_filter.ref = ws.dimensions  # Applique le filtre sur toute la plage de données

    # 2. Ajuste la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Numéro de colonne
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2  # +2 pour un petit confort visuel

    # 3. Souligne en bleu la dernière colonne si c'est un lien
    last_col_idx = ws.max_column
    for row in range(2, ws.max_row + 1):  # Ignore l'entête
        cell = ws.cell(row=row, column=last_col_idx)
        if isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.font = Font(color="0000FF", underline="single")

    # Sauvegarde finale
    wb.save("resultats_scenarios.xlsx")


    # Chargement du fichier
    wb = load_workbook("resultats_scenarios.xlsx")

    # Ajout d'une nouvelle feuille
    if "Details_Scenarios" in wb.sheetnames:
        ws = wb["Details_Scenarios"]
        wb.remove(ws)  # Nettoyage si déjà présent
    ws = wb.create_sheet("Details_Scenarios")

    # Style pour l'entête
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    row = 1
    for scenario in scenarios:
        meta = scenario[0]
        panneaux = scenario[1:]

        # Titre du scénario
        ws.cell(row=row, column=1, value=f"Scénario : {meta['libelleScenario']}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        # Infos globales
        ws.cell(row=row, column=1, value="Investissement (€)")
        ws.cell(row=row, column=2, value=meta["investissement"])
        ws.cell(row=row, column=3, value="Conso Maison (W)")
        ws.cell(row=row, column=4, value=meta["conso_maison_W"])
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = bold_font
        row += 2

        # En-têtes panneaux
        headers = ["Libellé", "Angle", "Aspect", "MaxPower", "TrackerType", "HorizonType"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = border
        row += 1

        # Infos par panneau
        for p in panneaux:
            ws.cell(row=row, column=1, value=p["libelle"])
            ws.cell(row=row, column=2, value=p["angle"])
            ws.cell(row=row, column=3, value=p["aspect"])
            ws.cell(row=row, column=4, value=p["MaxPower"])
            ws.cell(row=row, column=5, value=p["TrackerType"])
            ws.cell(row=row, column=6, value=p["horizonType"])
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
            row += 1

        row += 2  # Espace entre scénarios

 
#################

  
    ws_main = wb.worksheets[0]  # La première feuille (où sont listés les scénarios)

    # Création d'un dict pour accès rapide aux infos de scénario
    scenario_map = {}
    for scenario in scenarios:
        meta = scenario[0]
        panneaux = scenario[1:]
        texte =  f"Conso maison : {meta['conso_maison_W']} W"

        # Formatage des panneaux pour éviter trop de texte
        for i, p in enumerate(panneaux, 1):
            texte += f"\n• {p['libelle']}: {p['MaxPower']} kW, Angle {p['angle']}°, Aspect {p['aspect']}°, Track {p['TrackerType']}, Horizon {p['horizonType']} "
        
        scenario_map[meta['libelleScenario']] = texte

    # Appliquer les commentaires aux noms trouvés dans la colonne A
    for row in ws_main.iter_rows(min_row=2, min_col=1, max_col=1):
        cell = row[0]
        nom = str(cell.value)
        if nom in scenario_map:
            comment_text = scenario_map[nom]
            # Créer un commentaire
            comment = Comment(comment_text, "AutoGen")
            
            # Ajuster le commentaire (largeur et hauteur)
            comment.width = 300  # Largeur du commentaire (en pixels)
            comment.height = 150  # Hauteur du commentaire (en pixels)
            
            # Ajouter le commentaire à la cellule
            cell.comment = comment

  
    # Fonction pour convertir un numéro de colonne en lettre
    def column_index_to_letter(index):
        """Convertit un index de colonne (numérique) en lettre correspondante (ex: 1 -> 'A', 2 -> 'B')"""
        letter = ''
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter



    #wb = load_workbook("resultats_scenarios.xlsx")
    ws_main = wb.worksheets[0]  # La première feuille (où sont listés les scénarios)

    # Insertion de lignes en haut (avant la ligne 1)
    ws_main.insert_rows(1, 7)  # Insère 7 lignes au-dessus de la première ligne existante

    # Style pour les titres
    bold_font = Font(bold=True)

    # Ajout des données dans les nouvelles lignes
    ws_main.cell(row=1, column=1, value="Latitude")
    ws_main.cell(row=1, column=2, value=lat).font = bold_font

    ws_main.cell(row=2, column=1, value="Longitude")
    ws_main.cell(row=2, column=2, value=lon).font = bold_font

    ws_main.cell(row=3, column=1, value="Période")
    ws_main.cell(row=3, column=2, value=f"{startyear} - {endyear}").font = bold_font

    ws_main.cell(row=4, column=1, value="Prix kWh (€)")
    ws_main.cell(row=4, column=2, value=prix_kw).font = bold_font

    ws_main.cell(row=5, column=1, value="Date de départ cumul")
    ws_main.cell(row=5, column=2, value=date_departcumul.date()).font = bold_font

    ws_main.cell(row=6, column=1, value="Consommation chauffe-eau journalière (kWh)")
    ws_main.cell(row=6, column=2, value=conso_chauffeau_journalière).font = bold_font

    ws_main.cell(row=7, column=1, value="Consommation maison(W)")
    ws_main.cell(row=7, column=2, value=conso_maison_W).font = bold_font



    # Réappliquer le filtre à la nouvelle ligne 5 (au lieu de 1)
    last_column_letter = column_index_to_letter(ws_main.max_column)  # Convertir max_column en lettre
    ws_main.auto_filter.ref = f"A8:{last_column_letter}8"  # Définir la plage du filtre de la ligne 5

    # Sauvegarde du fichier
    wb.save(excelfilename)





def attendre_fermeture_fichier(fichier):
    while True:
        try:
            # On tente d'ouvrir en écriture exclusive
            with open(fichier, "a"):
                print(f"✅ Le fichier '{fichier}' est libre. On continue.")
                break
        except PermissionError:
            input(f"❌ Le fichier '{fichier}' est actuellement ouvert.\n👉 Merci de le fermer, puis appuyez sur [Entrée] pour réessayer...")
        time.sleep(1)



#----------------------------------------------
def calculate_solar_position(latitude, longitude, date):
    """Calcule la position solaire (azimut et élévation) pour une date et lieu donnés"""
    # Convertir en radians
    lat_rad = math.radians(latitude)
    
    # Jour de l'année
    day_of_year = date.timetuple().tm_yday
    
    # Déclinaison solaire
    delta = math.radians(23.45 * math.sin(math.radians((360.0/365.0) * (day_of_year - 81))))
    
    # Equation du temps (approximation)
    B = math.radians((360.0/365.0) * (day_of_year - 81))
    ET = 9.87 * math.sin(2*B) - 7.53 * math.cos(B) - 1.5 * math.sin(B)
    
    # Heure locale
    hour = date.hour + date.minute/60.0 + date.second/3600.0
    
    # Correction pour la longitude
    meridian = 15 * round(longitude/15)  # Méridien standard le plus proche
    LC = (meridian - longitude) / 15.0
    
    # Temps solaire
    solar_time = hour + ET/60.0 + LC
    
    # Angle horaire (en radians)
    hour_angle = math.radians(15.0 * (solar_time - 12.0))
    
    # Altitude (élévation)
    sin_altitude = (math.sin(lat_rad) * math.sin(delta) + 
                   math.cos(lat_rad) * math.cos(delta) * math.cos(hour_angle))
    altitude = math.degrees(math.asin(sin_altitude))
    
    # Azimut
    cos_azimuth = ((math.sin(delta) * math.cos(lat_rad) - 
                   math.cos(delta) * math.sin(lat_rad) * math.cos(hour_angle)) / 
                   math.cos(math.radians(altitude)))
    
    # Limiter cos_azimuth entre -1 et 1 pour éviter les erreurs numériques
    cos_azimuth = max(-1.0, min(1.0, cos_azimuth))
    
    azimuth = math.degrees(math.acos(cos_azimuth))
    
    # Ajuster l'azimut (0° au nord, sens horaire)
    if hour_angle > 0:
        azimuth = 360.0 - azimuth
    
    return azimuth, altitude

def get_sun_path_for_day(latitude, longitude, date):
    """Calcule la trajectoire du soleil pour une journée entière"""
    azimuths = []
    elevations = []
    
    # Calculer la position solaire toutes les 10 minutes
    for hour in range(24):
        for minute in range(0, 60, 10):
            current_time = date.replace(hour=hour, minute=minute, second=0)
            azimuth, elevation = calculate_solar_position(latitude, longitude, current_time)
            
            # Ne garder que les points où le soleil est au-dessus de l'horizon
            if elevation > 0:
                azimuths.append(azimuth)
                elevations.append(elevation)
    
    return azimuths, elevations

def create_solar_diagram(latitude, longitude, horizon_mask, save_path="diagramme_solaire.png"):
    """Crée et enregistre le diagramme solaire avec les trajectoires du soleil"""
    if len(horizon_mask) != 361:
        raise ValueError("Le masque d'horizon doit contenir 361 valeurs (azimut 0° à 360°).")
    
    # Configuration du graphique polaire
    fig, ax = plt.subplots(subplot_kw={'projection': 'polar'}, figsize=(10, 8))
    ax.set_theta_zero_location("N")
    ax.set_theta_direction(-1)
    
    # Tracer le masque d'horizon
    azimuths_deg = np.arange(0, 361)
    azimuths_rad = np.radians(azimuths_deg)
    elevation = 90 - np.array(horizon_mask)  # pour projection polaire
    ax.plot(azimuths_rad, elevation, label="Masque d'horizon", color="black", linewidth=2)
    
    # Dates pour chaque mois (21ème jour pour être proche des solstices/équinoxes)
    year = 2025  # Année actuelle
    dates = [datetime(year, month, 21) for month in range(1, 13)]
    
    # Couleurs pour chaque mois
    colors = plt.cm.hsv(np.linspace(0, 1, 12))
    
    # Tracer les trajectoires du soleil pour chaque mois
    for i, date in enumerate(dates):
        azimuths, elevations = get_sun_path_for_day(latitude, longitude, date)
        if azimuths:  # Vérifier qu'il y a des points à tracer
            # Pour l'affichage polaire
            radii = 90 - np.array(elevations)  
            theta = np.radians(azimuths)
            
            # Tracer la course du soleil
            line = ax.plot(theta, radii, color=colors[i], linewidth=1.5, 
                         label=date.strftime("%B"))
            
            # Marquer les heures (tous les 2 heures entre 6h et 18h)
            for hour in [6, 8, 10, 12, 14, 16, 18]:
                hour_date = date.replace(hour=hour, minute=0, second=0)
                azimuth, elevation = calculate_solar_position(latitude, longitude, hour_date)
                if elevation > 0:  # Si le soleil est visible
                    radius = 90 - elevation
                    ax.text(np.radians(azimuth), radius, f"{hour}", 
                           fontsize=8, ha='center', va='center', 
                           bbox=dict(facecolor='white', alpha=0.7, pad=1))
    
    # Configuration des limites et étiquettes
    ax.set_rlim(0, 90)
    ax.set_rticks([15, 30, 45, 60, 75, 90])
    ax.set_rlabel_position(135)
    ax.set_yticklabels(['75°', '60°', '45°', '30°', '15°', '0°'])  # Etiquettes d'élévation
    
    # Ajouter les étiquettes cardinales
    ax.set_xticklabels(['N', 'NE', 'E', 'SE', 'S', 'SO', 'O', 'NO'])
    
    # Titre et légende
    plt.title(f"Diagramme solaire avec trajectoires mensuelles\nLatitude: {latitude}°, Longitude: {longitude}°", 
             y=1.08)
    plt.legend(loc='upper center', bbox_to_anchor=(0.5, -0.05), ncol=4)
    
    # Ajouter une grille
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    
    # Enregistrer l'image
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    return save_path

def add_diagram_to_excel(excel_file,i,  image_path ,latitude , longitude ):
    """Ajoute le diagramme solaire à un nouvel onglet dans le fichier Excel"""
    # Charger le fichier Excel existant
    workbook = load_workbook(excel_file)
    
    # Créer un nouvel onglet pour le diagramme
    sheet_name = "Diagramme Solaire_"+str(i)
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)
    
    # Ajouter un titre
    worksheet['A1'] = "Diagramme Solaire - Course du soleil mois par mois"
    worksheet['A2'] = f"Coordonnées: Latitude {latitude}°, Longitude {longitude}°"
    
    # Insérer l'image
    img = Image(image_path)
    
    # Ajuster la taille de l'image si nécessaire (optionnel)
    scale_factor = 0.4  # Réduire à 75% de la taille originale
    img.width = img.width * scale_factor
    img.height = img.height * scale_factor
    
    # Positionner l'image sous le titre
    worksheet.add_image(img, 'A4')
    
    # Enregistrer le fichier Excel modifié
    workbook.save(excel_file)
    
    return True
        
#-------------------------------------------------------
#main



# Exemple d'utilisation
fichier = "resultats_scenarios.xlsx"
if os.path.exists(fichier):
    attendre_fermeture_fichier(fichier)


# Traitement de chaque scénario
for scenario in scenarios:
    meta = scenario[0]
    configs = scenario[1:]
   
    atleastone=doaall    #mettre 1 pour clean

    print(f"\n▶️ Scénario: {meta['libelleScenario']}")
    for config in configs:
        print(f"  - Panneau: angle={config['angle']}°, aspect={config['aspect']}°, power={config['MaxPower']} kW")

    df_cumul = pd.DataFrame()
    dfs_scenarios = {}

    nompourhash=""
    for i, params in enumerate(configs, 1):
        url, filename = build_pvgis_url(i,lat, lon, **params, startyear=startyear, endyear=endyear)
        updated = telecharger_csv(url, filename)

        nom_finalbrut = f"energie_brut_{params['libelle']}_{i}_{params['angle']}_{params['aspect']}_{params['MaxPower']}.xlsx"
        nompourhash=nompourhash+nom_finalbrut
        
        if updated == 1:
            atleastone=1
            # 🔄 Fichier mis à jour → retraitement du CSV
            brut = traitement_csv(filename, params["MaxPower"])
            brut["datetime"] = pd.to_datetime(brut["datetime"])
            brut = brut[["datetime", "p"]]
            brut.to_excel(nom_finalbrut,index=False)

        else:
            # 📂 Fichier déjà présent → relire Excel existant
            print("relecture",nom_finalbrut)
            brut = pd.read_excel(nom_finalbrut)
  
           # brut = brut[["datetime", "p"]]
             
        brut.set_index('datetime', inplace=True)
        dfs_scenarios[f"scenario_{i}"] = brut.copy()
        # ⚙️ Ajouter les puissances au cumul
        df_cumul = brut if df_cumul.empty else df_cumul.add(brut, fill_value=0)

    # Ajouter chaque scénario en colonne individuelle
    for name, df in dfs_scenarios.items():
        df_cumul[f"{name}"] = df["p"]

    # Export du fichier cumulatif brut
    thach= hashlib.sha256(nompourhash.encode()).hexdigest()[:8]    
    nom_scenario = f"energie_scenarios_{meta['libelleScenario']}_{thach}.xlsx"
    nom_final = f"energie_cumulee_scenarios_{meta['libelleScenario']}_{thach}.xlsx"

    if atleastone  or  not os.path.exists(nom_scenario) :
        print("regenere cumul pour ce scenario:",nom_scenario)
        df_cumul.to_excel(nom_scenario)

    # Traitement final
        df_cumulfin = traitement_cumul(nom_scenario, meta['conso_maison_W'])
        df_cumulfin.to_excel(nom_final)
        ajouter_graphique_excel(nom_final, df_cumulfin)
    
    else:
        df_cumulfin = pd.read_excel(nom_final)

   
    last_row = df_cumulfin[["cumul_besoin (Kwh)", "energie_perdue"	,"manque",	"energie_recuperee",  "cumul_gain (€)", "cumul_perte (€)" ,  "cumul_manque (€)" ]].iloc[-1]

    besoin=last_row["cumul_besoin (Kwh)"]
    gain = last_row["cumul_gain (€)"]
    perte = last_row["cumul_perte (€)"]
    manque =last_row["cumul_manque (€)"]
    
    investissement = meta["investissement"]

    ratio = perte / gain if gain else None
    gain_sur_invest = gain / investissement if investissement else None

    resultats.append({
        "scenario": meta["libelleScenario"],
        "investissement": investissement,
         "BesoinTotal(Kwh)": besoin,
        "gain(Kwh)": gain / prix_kw ,
        "perte(Kwh)":perte / prix_kw,
        "manque(Kwh)": manque / prix_kw,
        
        "BesoinTotal(€)": besoin * prix_kw ,
        "gain(€)": gain,
        "perte(€)": perte,
        "manque(€)": manque,
 
        "ratio_perte_gain": ratio,
        "gain_sur_invest": gain_sur_invest,
      "Depense": investissement+ manque,
      "Gain_Sur_depense": gain/ (investissement+ manque),


        
        "lien_excel": f'=HYPERLINK("{nom_final}", "Ouvrir Excel")'
    })


       

# Résumé final
print("\n📊 Résumé des scénarios :\n")
for res in resultats:
    print(f"🔹 {res['scenario']}")
    print(f"    Investissement     : {res['investissement']} €")
    print(f"    Gain cumulé        : {res['gain(€)']:.2f} €")
    print(f"    Perte cumulée      : {res['perte(€)']:.2f} €")
    print(f"    manque cumulée      : {res['manque(€)']:.2f} €")
    print(f"    Ratio perte/gain   : {res['ratio_perte_gain']:.2%}" if res['ratio_perte_gain'] is not None else "    ⚠️ Ratio perte/gain : gain = 0")
    print(f"    Gain / Invest      : {res['gain_sur_invest']:.2%}" if res['gain_sur_invest'] is not None else "    ⚠️ Impossible de calculer le rendement")
    print("")

# Export final des résultats
df_resultats = pd.DataFrame(resultats)

# 🔽 Appliquer la règle d’arrondi

def format_number(val):
        try:
            if isinstance(val, (int, float)):
                if abs(val) < 100:
                    return round(val,2)
                else:
                    return round(val)
            return val
        except:
            return val

df_resultats = df_resultats.applymap(format_number)

# Sauvegarde initiale avec pandas
saveto="resultats_scenarios.xlsx"
df_resultats.to_excel(saveto, index=False)
format_excel(saveto)

print("✅ Résumé global exporté dans 'resultats_scenarios.xlsx'")



for i in range(1, len(datahorizonType)):
    tableau = [float(x) for x in datahorizonType[i].split(',')]
    # Dupliquer le premier élément
    premier_element = tableau[0]
    tableau.insert(0, premier_element)  # Insérer une copie du premier élément au début
    # Créer et enregistrer le diagramme solaire
    diagram_path = create_solar_diagram(lat, lon, tableau)
    # Ajouter le diagramme à un nouvel onglet dans le fichier Excel
    add_diagram_to_excel(saveto, i, diagram_path, lat, lon)
    print(f"Le diagramme solaire pour l'horizon {i} a été ajouté avec succès au fichier Excel '{saveto}'.")
    

if os.path.exists(fichier):
    print(f"📂 Ouverture de {fichier}...")
    os.startfile(fichier)

